import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any, Optional
from datetime import datetime
from src.app.models.table import Table
from src.app.models.field import Field

class ExportService:
    
    @staticmethod
    async def export_to_excel(
        table: Table,
        records: List[Dict[str, Any]],
        include_aggregations: bool = False,
        aggregations: Dict = None
    ) -> StreamingResponse:
        """
        Экспорт данных таблицы в Excel
        """
        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table.name[:30]
        
        # Заголовки
        headers = [field.display_name for field in table.fields]
        ws.append(headers)
        
        # Стиль для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for record in records:
            row = []
            for field in table.fields:
                value = record.get(field.name, '')
                
                if field.field_type == 'number' and value:
                    row.append(float(value))
                elif field.field_type == 'date' and value:
                    row.append(value)
                elif field.field_type == 'boolean':
                    row.append('✓' if value else '✗')
                else:
                    row.append(str(value) if value else '')
            
            ws.append(row)
        
        # Итоги
        row_num = len(records) + 3
        if include_aggregations and aggregations:
            ws.cell(row=row_num, column=1, value="ИТОГО:").font = Font(bold=True)
            
            if aggregations.total_records:
                ws.cell(row=row_num + 1, column=1, value=f"Всего записей: {aggregations.total_records}")
            
            col_offset = 2
            for field_name, sum_value in aggregations.sums.items():
                ws.cell(row=row_num, column=col_offset, value=f"Сумма {field_name}:")
                ws.cell(row=row_num, column=col_offset + 1, value=sum_value)
                col_offset += 2
        
        # Автоширина
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Возвращаем
        safe_filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={safe_filename}"
            }
        )